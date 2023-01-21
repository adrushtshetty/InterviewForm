from flask import Flask, request, jsonify, render_template, url_for, redirect
from csv import writer
import pandas as pd
app = Flask(__name__)


@app.route('/', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form['username'] != 'androengrsupt' or request.form['password'] != 'cnooc114':
            error = 'Invalid Credentials. Please try again.'
        else:
            return redirect(url_for('home_page'))
    return render_template('login.html', error=error)

@app.route('/home')
def home_page():
    return render_template("index.html")

@app.route('/interview_form')
def interview():
    return render_template("forms-elements.html")

@app.route('/trials')
def trial():
    return render_template("interviewENG.html")

@app.route('/trials',methods=['POST'])
def enter():
    # import shutil
    # nm="F - IMS - 160 Interview Form " + request.form['Name'] + " " + request.form['Rank']
    # filename1 = nm + '.xlsx'
    # shutil.copy('template.xlsx', filename1)
    # from openpyxl import load_workbook
    # workbook = load_workbook(filename=filename1)
    # sheet = workbook.active
    import shutil
    nm="F-IMS-160 Interview Form " + request.form['Name'] + " " + request.form['Rank']
    filename1 = "C:\\Users\\Admin\\Desktop\\InterviewForms\\" + nm + '.xlsx'
    shutil.copy('C:\\Users\\Admin\\Desktop\\InterviewForms\\Template.xlsx', filename1)

    from openpyxl import load_workbook
    workbook = load_workbook(filename=filename1)
    sheet = workbook.active
    sheet["C6"] = request.form['Name']
    sheet["F6"] = request.form['Department']
    sheet["H6"] = request.form['Rank']
    sheet["J6"] = request.form['Date']
    sheet["D12"] = request.form['Age']
    sheet["F12"] = request.form['RankXP']
    sheet["J12"] = request.form['CES']
    sheet["F14"] = request.form['EXPtanker']
    sheet["H12"] = request.form['VesselXP']
    sheet["J14"] = request.form['IceCond']
    sheet["D15"] = request.form['Fourth']
    sheet["F15"] = request.form['Third']
    sheet["H15"] = request.form['Second']
    sheet["J15"] = request.form['Chief']
    sheet["D17"] = request.form['engineRoom']
    sheet["H17"] = request.form['H17']
    sheet["I17"] = request.form['I17']
    sheet["E18"] = request.form['E18']
    sheet["J29"] = request.form['J29']
    sheet["D31"] = request.form['D31']
    sheet["F31"] = request.form['F31']
    sheet["H31"] = request.form['H31']
    sheet["H31"] = request.form['H31']
    sheet["J31"] = request.form['J31']
    sheet["C33"] = request.form['C33']
    sheet["D33"] = request.form['D33']
    sheet["G33"] = request.form['G33']
    sheet["H33"] = request.form['H33']
    sheet["D8"] = request.form['D8']
    sheet["F8"] = request.form['F8']

    sheet["H18"] = request.form['H18']
    sheet["J18"] = request.form['J18']
    sheet["D19"] = request.form['D19']
    sheet["G19"] = request.form['G19']
    sheet["I19"] = request.form['I19']
    sheet["J19"] = request.form['J19']
    sheet["D21"] = request.form['D21']
    sheet["F21"] = request.form['F21']
    sheet["H21"] = request.form['H21']
    sheet["J21"] = request.form['J21']
    sheet["D22"] = request.form['D22']
    sheet["F22"] = request.form['F22']
    sheet["H22"] = request.form['H22']
    sheet["J22"] = request.form['J22']
    sheet["D23"] = request.form['D23']
    sheet["F23"] = request.form['F23']
    sheet["H23"] = request.form['H23']
    sheet["J23"] = request.form['J23']
    #Ends here
    # sheet["D25"] = request.form['D25']
    # sheet["F25"] = request.form['F25']
    # sheet["H25"] = request.form['H25']
    # sheet["J25"] = request.form['J25']
    # sheet["D26"] = request.form['D26']
    # sheet["F26"] = request.form['F26']
    # sheet["H26"] = request.form['H26']
    # sheet["J26"] = request.form['J26']
    # sheet["D28"] = request.form['D28']
    # sheet["F28"] = request.form['F28']
    # sheet["H28"] = request.form['H28']
    # sheet["J28"] = request.form['J28']
    # sheet["D29"] = request.form['D29']
    # sheet["F29"] = request.form['F29']
    # sheet["H29"] = request.form['H29']
    #
    # sheet["J29"] = request.form['J29']
    # sheet["D31"]=request.form['D31']
    # sheet["F31"] = request.form['F31']
    # sheet["H31"] = request.form['H31']
    # sheet["J31"] = request.form['J31']
    # sheet["C33"] = request.form['C33']
    # sheet["D33"] = request.form['D33']
    # sheet["G33"] = request.form['G33']
    # sheet["H33"] = request.form['H33']
    # sheet["D8"] = request.form['D8']
    # sheet["F8"] = request.form['F8']
    #

    workbook.save(filename=filename1)
    return render_template("interviewENG.html")


@app.route('/eo_form')
def query_page():
    return render_template('eo_Form.html')


if __name__ == '__main__':
    app.run(debug=True)